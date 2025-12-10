from rest_framework import generics, status, serializers
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.decorators import (
    api_view,
    permission_classes,
    authentication_classes,
)
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

from django.contrib.auth import get_user_model
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse

from datetime import timedelta
from io import BytesIO
from collections import defaultdict
import secrets

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from .serializers import (
    UserRegisterSerializer,
    LessonProgressSerializer,
    generate_secure_password,
)
from .models import CustomUser, LessonProgress, Course

User = get_user_model()


# ========================================================
# 🔐 HELPER: generar y enviar código de verificación
# ========================================================
def _generate_and_send_verification_code(user: CustomUser):
    """
    Genera un código de 6 dígitos, lo guarda con expiración
    y lo envía al usuario por correo.
    """
    code = f"{secrets.randbelow(1000000):06d}"

    user.verification_code = code
    user.verification_code_expires_at = timezone.now() + timedelta(minutes=15)
    user.save(update_fields=["verification_code", "verification_code_expires_at"])

    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@cyberkids.cl")

    send_mail(
        subject="Código de verificación - CyberKids Chile",
        message=(
            f"Hola {user.username},\n\n"
            f"Tu código de verificación es: {code}\n"
            f"Este código expira en 15 minutos.\n\n"
            f"Equipo CyberKids Chile"
        ),
        from_email=from_email,
        recipient_list=[user.email],
        fail_silently=False,
    )


# ========================================================
# 🔐 HELPER: validar contraseña segura (reusa el serializer)
# ========================================================
def _validate_secure_password(password: str):
    """
    Reutiliza la misma lógica de validación que el registro
    (UserRegisterSerializer.validate_password).
    Lanza serializers.ValidationError si no cumple.
    """
    serializer = UserRegisterSerializer()
    return serializer.validate_password(password)


# ========================================================
# 🔵 HELPER: stats de estudiante (reutilizable para docente)
# ========================================================
def _get_student_stats(student: CustomUser):
    """
    Calcula XP total y stats básicos de un estudiante.
    Usado en panel docente.
    """
    qs = LessonProgress.objects.filter(user=student)
    total_xp = 0
    total_score = 0.0
    total_time = 0.0
    n = qs.count()

    for p in qs:
        total_xp += p.xp
        total_score += p.score
        total_time += p.time

    avg_score = total_score / n if n else 0
    avg_time = total_time / n if n else 0

    return {
        "id": student.id,
        "username": student.username,
        "age": student.age,
        "age_group": student.age_group,
        "total_xp": total_xp,
        "average_score": avg_score,
        "average_time": avg_time,
    }


# ========================================================
# 🟦 REGISTRO
# ========================================================
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegisterSerializer
    permission_classes = [AllowAny]
    authentication_classes = []

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()

            # Solo ADULTOS deben verificar
            if user.role in ["parent", "teacher", "school", "admin"]:
                _generate_and_send_verification_code(user)

            child_password = getattr(user, "generated_child_password", None)

            return Response(
                {
                    "message": "Usuario creado correctamente. Revisa tu correo para verificar la cuenta.",
                    "username": user.username,
                    "email": user.email,
                    "role": user.role,
                    "child_password": child_password,
                },
                status=201,
            )

        return Response(serializer.errors, status=400)


# ========================================================
# 🟦 LOGIN (con email/usuario, verificación y archivado)
# ========================================================
class EmailTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        username_or_email = attrs.get("username")
        expected_role = self.context["request"].data.get("expected_role")

        # Permite usar email como username → reemplazamos en attrs
        if username_or_email:
            try:
                user_obj = User.objects.get(email__iexact=username_or_email)
                attrs["username"] = user_obj.username
            except User.DoesNotExist:
                # Si no existe ese email, se intenta como username normal
                pass

        # Autenticación estándar de SimpleJWT
        data = super().validate(attrs)
        user = self.user  # definido por la clase base

        if not user.is_active:
            raise serializers.ValidationError("Cuenta inactiva.")

        # 🚫 Bloqueo si el usuario fue archivado (soft delete)
        if getattr(user, "is_archived", False):
            raise serializers.ValidationError(
                "Esta cuenta ha sido desactivada por el apoderado."
            )

        # ❗ Bloqueo si NO ha verificado (adultos)
        if user.role in ["parent", "teacher", "school", "admin"] and not user.email_verified:
            raise serializers.ValidationError(
                "Debes verificar tu correo antes de iniciar sesión."
            )

        # Validar que el rol esperado coincida
        if expected_role and user.role != expected_role:
            raise serializers.ValidationError(
                f"No tienes permisos para ingresar como {expected_role}."
            )

        # Añadimos datos extra al payload
        data["username"] = user.username
        data["role"] = user.role
        return data


class EmailTokenObtainPairView(TokenObtainPairView):
    serializer_class = EmailTokenObtainPairSerializer


# ========================================================
# 🟦 ENVIAR / REENVIAR CÓDIGO DE VERIFICACIÓN
# ========================================================
@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def send_verification_code(request):
    email = request.data.get("email")
    if not email:
        return Response({"error": "Debes enviar el correo."}, status=400)

    email = email.lower().strip()

    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({"error": "No existe un usuario con este correo."}, status=404)

    if user.email_verified:
        return Response({"message": "Este correo ya está verificado."}, status=200)

    _generate_and_send_verification_code(user)

    return Response({"message": "Código de verificación enviado."}, status=200)


# ========================================================
# 🟦 VERIFICAR CORREO
# ========================================================
@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def verify_email(request):
    email = request.data.get("email")
    code = request.data.get("code")

    if not email or not code:
        return Response({"error": "Debes enviar correo y código."}, status=400)

    email = email.lower().strip()

    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({"error": "No existe un usuario con este correo."}, status=404)

    if user.email_verified:
        return Response({"message": "El correo ya está verificado."}, status=200)

    # Validar código existente
    if not user.verification_code:
        return Response(
            {"error": "No hay código activo. Solicita uno nuevo."}, status=400
        )

    if timezone.now() > user.verification_code_expires_at:
        return Response({"error": "El código ha expirado."}, status=400)

    if code != user.verification_code:
        return Response({"error": "Código incorrecto."}, status=400)

    # Marcar como verificado
    user.email_verified = True
    user.verification_code = None
    user.verification_code_expires_at = None
    user.save(
        update_fields=[
            "email_verified",
            "verification_code",
            "verification_code_expires_at",
        ]
    )

    return Response({"message": "Correo verificado correctamente."}, status=200)


# ========================================================
# 🟦 OBTENER ESTUDIANTE POR USERNAME (no archivados)
# ========================================================
@api_view(["GET"])
def get_student_by_username(request, username):
    try:
        student = CustomUser.objects.get(
            username=username, role="student", is_archived=False
        )
    except CustomUser.DoesNotExist:
        return Response({"error": "Estudiante no encontrado"}, status=404)

    return Response(
        {
            "id": student.id,
            "username": student.username,
            "email": student.email,
            "linked_parent": student.linked_student.username
            if student.linked_student
            else None,
            "age_group": student.age_group,
            "age": student.age,
        }
    )


# ========================================================
# 🟦 SET AGE GROUP
# ========================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def set_student_age_group(request):
    username = request.data.get("username")
    age_group = request.data.get("age_group")

    if not username or not age_group:
        return Response({"error": "Faltan campos"}, status=400)

    try:
        student = CustomUser.objects.get(
            username=username, role="student", is_archived=False
        )
    except CustomUser.DoesNotExist:
        return Response({"error": "Estudiante no encontrado"}, status=404)

    student.age_group = age_group
    student.save()
    return Response({"message": "Grupo de edad actualizado correctamente"})


# ========================================================
# 🟦 PROGRESO
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_my_progress(request):
    progress = LessonProgress.objects.filter(user=request.user)
    return Response(LessonProgressSerializer(progress, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_lesson_progress(request):
    lesson_id = request.data.get("lesson_id")

    if not lesson_id:
        return Response({"error": "lesson_id requerido"}, status=400)

    progress, created = LessonProgress.objects.update_or_create(
        user=request.user,
        lesson_id=lesson_id,
        defaults={
            "score": request.data.get("score", 1),
            "time": request.data.get("time", 0),
            "xp": request.data.get("xp", 0),
            "completed": True,
        },
    )

    return Response(LessonProgressSerializer(progress).data)


# ========================================================
# 🟦 LEADERBOARD (ignora alumnos archivados)
# ========================================================
@api_view(["GET"])
def get_leaderboard(request):
    leaderboard = []

    for student in CustomUser.objects.filter(role="student", is_archived=False):
        total_xp = sum(p.xp for p in LessonProgress.objects.filter(user=student))
        leaderboard.append(
            {
                "username": student.username,
                "xp": total_xp,
                "age_group": student.age_group,
            }
        )

    leaderboard.sort(key=lambda x: x["xp"], reverse=True)

    return Response(leaderboard)


# ========================================================
# 🟦 MI PERFIL
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_me(request):
    user = request.user
    return Response(
        {
            "username": user.username,
            "email": user.email,
            "role": user.role,
            "age_group": user.age_group,
        }
    )


# ========================================================
# 🟥 PADRES — GESTIÓN DE HIJOS
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_children(request):
    parent = request.user
    children = CustomUser.objects.filter(
        linked_student=parent, role="student", is_archived=False
    )

    return Response(
        [
            {
                "id": c.id,
                "username": c.username,
                "age": c.age,
                "age_group": c.age_group,
                "password": "",
                "password_changed_once": c.password_changed_once,
            }
            for c in children
        ]
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_child_api(request):
    parent = request.user
    username = request.data.get("username")
    age = request.data.get("age")

    if not username:
        return Response({"error": "Nombre de usuario obligatorio"}, status=400)

    if CustomUser.objects.filter(username=username).exists():
        return Response({"error": "Ese usuario ya existe"}, status=400)

    password = generate_secure_password()
    fake_email = f"{username.lower()}@child.cyberkids.cl"

    CustomUser.objects.create_user(
        username=username,
        email=fake_email,
        password=password,
        role="student",
        age=age,
        linked_student=parent,
    )

    return Response(
        {
            "message": "Hijo creado correctamente",
            "child_username": username,
            "child_password": password,
        },
        status=201,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_child(request):
    child_id = request.data.get("child_id")

    try:
        child = CustomUser.objects.get(
            id=child_id,
            linked_student=request.user,
            role="student",
            is_archived=False,
        )
    except CustomUser.DoesNotExist:
        return Response({"error": "Hijo no encontrado."}, status=404)

    new_username = request.data.get("new_username")
    new_password = request.data.get("new_password")
    old_password = request.data.get("old_password")

    # Validar username duplicado
    if new_username:
        if (
            CustomUser.objects.filter(username=new_username)
            .exclude(id=child_id)
            .exists()
        ):
            return Response(
                {"error": "Nombre de usuario ya registrado."}, status=400
            )
        child.username = new_username

    # Cambiar contraseña
    if new_password:

        if child.password_changed_once:
            if not old_password:
                return Response(
                    {"error": "Debes ingresar la contraseña anterior."},
                    status=400,
                )

            if not child.check_password(old_password):
                return Response(
                    {"error": "La contraseña anterior no es correcta."},
                    status=400,
                )

        child.set_password(new_password)
        child.password_changed_once = True

    child.save()

    return Response({"message": "Hijo actualizado correctamente."})


# ========================================================
# 🟥 PADRES — ARCHIVAR (soft delete) HIJO
# ========================================================
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def archive_child(request):
    child_id = request.data.get("child_id")

    if not child_id:
        return Response({"error": "child_id requerido."}, status=400)

    try:
        child = CustomUser.objects.get(
            id=child_id,
            linked_student=request.user,
            role="student",
            is_archived=False,
        )
    except CustomUser.DoesNotExist:
        return Response({"error": "Hijo no encontrado."}, status=404)

    child.is_archived = True
    child.save(update_fields=["is_archived"])

    return Response({"message": "Alumno archivado correctamente."})


# ========================================================
# 🟥 PADRES — PROGRESO ESPECÍFICO DE UN HIJO (no archivados)
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_child_progress_for_parent(request, child_id):
    parent = request.user
    try:
        child = CustomUser.objects.get(
            id=child_id,
            linked_student=parent,
            role="student",
            is_archived=False,
        )
    except CustomUser.DoesNotExist:
        return Response(
            {"error": "Hijo no encontrado o no pertenece a este apoderado."},
            status=404,
        )

    progress_qs = LessonProgress.objects.filter(user=child)
    progress_serializer = LessonProgressSerializer(progress_qs, many=True)

    return Response(
        {
            "id": child.id,
            "username": child.username,
            "age_group": child.age_group,
            "age": child.age,
            "progress": progress_serializer.data,
        }
    )


# ========================================================
# 🟥 ADMIN — LISTAR / CREAR USUARIOS
# ========================================================
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def admin_list_users(request):
    # solo admins
    if request.user.role != "admin":
        return Response({"error": "No autorizado"}, status=403)

    # ---------------- GET: listar con paginación ----------------
    if request.method == "GET":
        only_active = (
            request.query_params.get("only_active", "true").lower() == "true"
        )
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 20))

        qs = CustomUser.objects.all().order_by("-date_joined")

        if only_active:
            qs = qs.filter(is_archived=False)

        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        qs = qs[start:end]

        data = [
            {
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "role": u.role,
                "age": u.age,
                "age_group": u.age_group,
                "linked_student_username": (
                    u.linked_student.username if u.linked_student else None
                ),
                "is_archived": u.is_archived,
                "date_joined": u.date_joined,
                "last_login": u.last_login,
            }
            for u in qs
        ]

        return Response(
            {"results": data, "page": page, "page_size": page_size, "total": total}
        )

    # ---------------- POST: crear usuario ----------------
    if request.method == "POST":
        username = request.data.get("username")
        password = request.data.get("password")
        role = request.data.get("role")
        email = request.data.get("email")

        if not username or not password or not role:
            return Response(
                {"error": "username, password y role son obligatorios."},
                status=400,
            )

        if role not in ["student", "parent", "teacher", "school", "admin"]:
            return Response(
                {
                    "error": "Rol inválido. Usa student, parent, teacher, school o admin."
                },
                status=400,
            )

        if CustomUser.objects.filter(username=username).exists():
            return Response({"error": "Ese nombre de usuario ya existe."}, status=400)

        # ⚠️ Validación de contraseña segura (misma que registro)
        try:
            _validate_secure_password(password)
        except serializers.ValidationError as e:
            # e.detail puede ser string o list
            msg = str(e.detail if not isinstance(e.detail, list) else e.detail[0])
            return Response({"error": msg}, status=400)

        # Si no se envía email, generamos uno falso
        if not email:
            email = f"{username.lower()}@admin-created.cyberkids.local"

        # Normalizamos email
        email = email.lower().strip()

        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            role=role,
        )

        # ✅ Usuarios creados desde el panel admin se consideran verificados
        user.email_verified = True
        user.save(update_fields=["email_verified"])

        return Response(
            {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "role": user.role,
                "is_archived": user.is_archived,
            },
            status=201,
        )


# ========================================================
# 🟥 ADMIN — EDITAR / ARCHIVAR USUARIO
# ========================================================
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def admin_update_user(request, user_id):
    # solo admins
    if request.user.role != "admin":
        return Response({"error": "No autorizado"}, status=403)

    try:
        user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
        return Response({"error": "Usuario no encontrado."}, status=404)

    new_username = request.data.get("username")
    new_role = request.data.get("role")
    new_password = request.data.get("password")
    is_archived = request.data.get("is_archived", None)

    # Cambiar username
    if new_username:
        if (
            CustomUser.objects.filter(username=new_username)
            .exclude(id=user.id)
            .exists()
        ):
            return Response(
                {"error": "Ese nombre de usuario ya existe."}, status=400
            )
        user.username = new_username

    # Cambiar rol
    if new_role:
        if new_role not in ["student", "parent", "teacher", "school", "admin"]:
            return Response(
                {
                    "error": "Rol inválido. Usa student, parent, teacher, school o admin."
                },
                status=400,
            )
        user.role = new_role

    # Cambiar contraseña
    if new_password:
        try:
            _validate_secure_password(new_password)
        except serializers.ValidationError as e:
            msg = str(e.detail if not isinstance(e.detail, list) else e.detail[0])
            return Response({"error": msg}, status=400)

        user.set_password(new_password)

    # Archivar / desarchivar
    if is_archived is not None:
        # Se espera boolean o algo convertible
        if isinstance(is_archived, str):
            is_archived = is_archived.lower() == "true"
        user.is_archived = bool(is_archived)

    user.save()

    return Response(
        {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.role,
            "is_archived": user.is_archived,
        }
    )


# ========================================================
# 🟥 ADMIN — RESUMEN ESTADÍSTICAS (JSON para gráficos)
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_report_overview(request):
    if request.user.role != "admin":
        return Response({"error": "No autorizado"}, status=403)

    # Solo alumnos activos
    students_qs = CustomUser.objects.filter(role="student", is_archived=False)
    progress_qs = LessonProgress.objects.filter(
        user__role="student",
        user__is_archived=False,
    )

    total_students = students_qs.count()
    archived_students = CustomUser.objects.filter(
        role="student", is_archived=True
    ).count()
    total_lessons_completed = progress_qs.count()

    # promedio global
    total_score = 0.0
    total_time = 0.0
    for p in progress_qs:
        total_score += p.score
        total_time += p.time

    avg_score = (
        total_score / total_lessons_completed if total_lessons_completed else 0
    )
    avg_time = total_time / total_lessons_completed if total_lessons_completed else 0

    # ----------------- por lección -----------------
    per_lesson = defaultdict(
        lambda: {
            "lesson_id": "",
            "times_played": 0,
            "sum_score": 0.0,
            "sum_time": 0.0,
            "errors": 0,  # intentos con score < 1
            "sum_xp": 0,
        }
    )

    for p in progress_qs:
        d = per_lesson[p.lesson_id]
        d["lesson_id"] = p.lesson_id
        d["times_played"] += 1
        d["sum_score"] += p.score
        d["sum_time"] += p.time
        d["sum_xp"] += p.xp
        if p.score < 1.0:
            d["errors"] += 1

    per_lesson_list = []
    for lesson_id, d in per_lesson.items():
        times = d["times_played"] or 1
        error_rate = (d["errors"] / times) * 100.0
        per_lesson_list.append(
            {
                "lesson_id": lesson_id,
                "times_played": d["times_played"],
                "average_score": d["sum_score"] / times,
                "average_time": d["sum_time"] / times,
                "error_rate": error_rate,  # porcentaje
                "total_xp": d["sum_xp"],
            }
        )

    # Top juegos por errores (mayor % error)
    top_lessons_by_errors = sorted(
        per_lesson_list, key=lambda x: x["error_rate"], reverse=True
    )[:5]

    # Top alumnos por XP (usando el mismo criterio del leaderboard)
    top_students = []
    for s in students_qs:
        total_xp = sum(p.xp for p in LessonProgress.objects.filter(user=s))
        top_students.append(
            {
                "username": s.username,
                "xp": total_xp,
                "age_group": s.age_group,
            }
        )

    top_students.sort(key=lambda x: x["xp"], reverse=True)
    top_students = top_students[:5]

    data = {
        "total_students": total_students,
        "archived_students": archived_students,
        "total_lessons_completed": total_lessons_completed,
        "average_score": avg_score,
        "average_time": avg_time,
        "per_lessons": per_lesson_list,
        "top_lessons_by_errors": top_lessons_by_errors,
        "top_students": top_students,
    }

    return Response(data)


# ========================================================
# 🟥 ADMIN — EXPORTAR REPORTE (Excel / PDF)
# ========================================================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_report_export(request):
    if request.user.role != "admin":
        return Response({"error": "No autorizado"}, status=403)

    export_format = request.query_params.get("format", "excel").lower()

    # reutilizamos la lógica de overview
    overview_response = admin_report_overview(request)
    if overview_response.status_code != 200:
        return overview_response

    data = overview_response.data
    per_lessons = data.get("per_lessons", [])
    top_students = data.get("top_students", [])

    if export_format == "excel":
        # ---------- EXCEL ----------
        wb = Workbook()

        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"

        ws1["A1"] = "Total alumnos activos"
        ws1["B1"] = data["total_students"]

        ws1["A2"] = "Alumnos archivados"
        ws1["B2"] = data["archived_students"]

        ws1["A3"] = "Lecciones completadas"
        ws1["B3"] = data["total_lessons_completed"]

        ws1["A4"] = "Puntaje promedio global"
        ws1["B4"] = data["average_score"]

        ws1["A5"] = "Tiempo promedio global (s)"
        ws1["B5"] = data["average_time"]

        # Hoja 2: Juegos
        ws2 = wb.create_sheet(title="Juegos")
        headers = [
            "Lesson ID",
            "Veces jugado",
            "Puntaje promedio",
            "Tiempo promedio (s)",
            "% intentos con error",
            "XP total",
        ]
        ws2.append(headers)

        for l in per_lessons:
            ws2.append(
                [
                    l["lesson_id"],
                    l["times_played"],
                    l["average_score"],
                    l["average_time"],
                    l["error_rate"],
                    l["total_xp"],
                ]
            )

        # Hoja 3: Top alumnos
        ws3 = wb.create_sheet(title="Top alumnos")
        ws3.append(["Usuario", "XP total", "Grupo etario"])
        for s in top_students:
            ws3.append([s["username"], s["xp"], s["age_group"]])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="reporte_cyberkids.xlsx"'
        )
        return response

    elif export_format == "pdf":
        # ---------- PDF ----------
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        y = height - 50
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Reporte CyberKids - Resumen")
        y -= 30

        p.setFont("Helvetica", 10)
        lines = [
            f"Total alumnos activos: {data['total_students']}",
            f"Alumnos archivados: {data['archived_students']}",
            f"Lecciones completadas: {data['total_lessons_completed']}",
            f"Puntaje promedio global: {data['average_score']:.2f}",
            f"Tiempo promedio global (s): {data['average_time']:.2f}",
            "",
            "Top juegos con mayor tasa de error:",
        ]

        for line in lines:
            p.drawString(50, y, line)
            y -= 15

        for l in data["top_lessons_by_errors"]:
            if y < 80:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 10)
            p.drawString(
                60,
                y,
                f"- {l['lesson_id']}: {l['error_rate']:.1f}% errores, "
                f"{l['times_played']} intentos",
            )
            y -= 15

        y -= 20
        p.drawString(50, y, "Top alumnos por XP:")
        y -= 15

        for s in top_students:
            if y < 80:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 10)
            p.drawString(
                60,
                y,
                f"- {s['username']}: {s['xp']} XP (grupo: {s['age_group']})",
            )
            y -= 15

        p.showPage()
        p.save()

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="reporte_cyberkids.pdf"'
        )
        return response

    else:
        return Response(
            {"error": "Formato no soportado. Usa excel o pdf."}, status=400
        )


# ========================================================
# 👨‍🏫 DOCENTE — CURSOS Y ALUMNOS
# ========================================================
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def teacher_list_courses(request):
    """
    GET: lista los cursos del docente logueado.
    POST: crea un nuevo curso para el docente.
    """
    user = request.user

    if user.role != "teacher":
        return Response({"error": "No autorizado"}, status=403)

    # --------- GET: listar cursos ----------
    if request.method == "GET":
        courses = Course.objects.filter(teacher=user).order_by("name")

        data = []
        for c in courses:
            students_qs = c.students.filter(role="student", is_archived=False)
            data.append(
                {
                    "id": c.id,
                    "name": c.name,
                    "grade": c.grade,
                    "students_count": students_qs.count(),
                }
            )

        return Response(data)

    # --------- POST: crear curso ----------
    if request.method == "POST":
        name = request.data.get("name")
        grade = request.data.get("grade")

        if not name:
            return Response(
                {"error": "El nombre del curso es obligatorio."}, status=400
            )

        course = Course.objects.create(
            name=name,
            grade=grade or None,
            teacher=user,
        )

        return Response(
            {
                "id": course.id,
                "name": course.name,
                "grade": course.grade,
                "students_count": 0,
            },
            status=201,
        )


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def teacher_course_detail(request, course_id):
    """
    GET   → detalle del curso + alumnos
    PATCH → actualizar nombre / grado del curso
    DELETE → eliminar el curso
    """
    user = request.user

    if user.role != "teacher":
        return Response({"error": "No autorizado"}, status=403)

    try:
        course = Course.objects.get(id=course_id, teacher=user)
    except Course.DoesNotExist:
        return Response({"error": "Curso no encontrado"}, status=404)

    # ---------- GET ----------
    if request.method == "GET":
        students_qs = course.students.filter(role="student", is_archived=False)
        students_data = [_get_student_stats(s) for s in students_qs]

        return Response(
            {
                "id": course.id,
                "name": course.name,
                "grade": course.grade,
                "students": students_data,
            }
        )

    # ---------- PATCH (editar curso) ----------
    if request.method == "PATCH":
        name = request.data.get("name")
        grade = request.data.get("grade", None)

        if name:
            course.name = name
        course.grade = grade if grade is not None else None
        course.save()

        students_qs = course.students.filter(role="student", is_archived=False)
        students_data = [_get_student_stats(s) for s in students_qs]

        return Response(
            {
                "id": course.id,
                "name": course.name,
                "grade": course.grade,
                "students": students_data,
            }
        )

    # ---------- DELETE (eliminar curso) ----------
    if request.method == "DELETE":
        course.delete()
        return Response(status=204)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def teacher_add_student_to_course(request, course_id):
    """
    Crea un alumno y lo agrega al curso del docente.
    Devuelve username + contraseña generada.
    """
    teacher = request.user

    if teacher.role != "teacher":
        return Response({"error": "No autorizado"}, status=403)

    try:
        course = Course.objects.get(id=course_id, teacher=teacher)
    except Course.DoesNotExist:
        return Response({"error": "Curso no encontrado"}, status=404)

    username = request.data.get("username")
    age = request.data.get("age")

    if not username:
        return Response({"error": "Nombre de usuario obligatorio"}, status=400)

    if CustomUser.objects.filter(username=username).exists():
        return Response({"error": "Ese usuario ya existe"}, status=400)

    # Generamos contraseña segura reutilizando tu helper
    password = generate_secure_password()
    fake_email = f"{username.lower()}@student.cyberkids.cl"

    student = CustomUser.objects.create_user(
        username=username,
        email=fake_email,
        password=password,
        role="student",
        age=age,
    )

    # Lo agregamos al curso
    course.students.add(student)

    return Response(
        {
            "message": "Alumno creado y agregado al curso correctamente",
            "student_username": student.username,
            "student_password": password,
            "course_id": course.id,
        },
        status=201,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def teacher_remove_student_from_course(request, course_id):
    """
    Quita un alumno (student_id) de un curso del docente.
    No elimina la cuenta del alumno, solo la relación.
    """
    teacher = request.user

    if teacher.role != "teacher":
        return Response({"error": "No autorizado"}, status=403)

    try:
        course = Course.objects.get(id=course_id, teacher=teacher)
    except Course.DoesNotExist:
        return Response({"error": "Curso no encontrado"}, status=404)

    student_id = request.data.get("student_id")
    if not student_id:
        return Response({"error": "student_id requerido"}, status=400)

    try:
        student = CustomUser.objects.get(id=student_id, role="student")
    except CustomUser.DoesNotExist:
        return Response({"error": "Alumno no encontrado"}, status=404)

    course.students.remove(student)

    return Response(
        {
            "message": "Alumno quitado del curso correctamente.",
            "course_id": course.id,
            "student_id": student.id,
        },
        status=200,
    )
